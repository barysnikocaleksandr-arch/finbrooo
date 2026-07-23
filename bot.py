import asyncio
import re
import sqlite3
import io
import os
import pandas as pd
import matplotlib.pyplot as plt
from aiogram import Bot, Dispatcher, types
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardMarkup, InlineKeyboardButton
from datetime import datetime, timedelta
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger

# --- Проверка библиотек ---
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import BarChart, Reference
except ImportError:
    pass

# --- Заглушка для Render ---
from aiohttp import web
import threading
async def handle(request):
    return web.Response(text="Bot is running!")
def run_web_server():
    app = web.Application()
    app.router.add_get('/', handle)
    web.run_app(app, host='0.0.0.0', port=10000)
threading.Thread(target=run_web_server, daemon=True).start()
# ---------------------------

BOT_TOKEN = "8856832421:AAEWvsUoVd5XTpOsnRcSfWSrCsM8jlvp-mw"

bot = Bot(token=BOT_TOKEN)
dp = Dispatcher()

# --- БАЗА ДАННЫХ ---
DB_PATH = '/tmp/finance.db'
conn = sqlite3.connect(DB_PATH)
cursor = conn.cursor()
cursor.execute('''
    CREATE TABLE IF NOT EXISTS transactions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER,
        type TEXT,
        amount INTEGER,
        category TEXT,
        date TEXT,
        comment TEXT,
        mood TEXT
    )
''')
cursor.execute('''
    CREATE TABLE IF NOT EXISTS user_limits (
        user_id INTEGER PRIMARY KEY,
        daily_limit INTEGER
    )
''')
conn.commit()

# --- КНОПКИ МЕНЮ ---
def get_main_keyboard():
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="💰 Баланс"), KeyboardButton(text="📊 Статистика")],
            [KeyboardButton(text="⚙️ Мой лимит"), KeyboardButton(text="🎯 Цель")],
            [KeyboardButton(text="🧘 Тренер"), KeyboardButton(text="📁 Скачать Excel")]
        ],
        resize_keyboard=True
    )

def get_mood_keyboard():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="😊 Классно", callback_data="mood_good"),
         InlineKeyboardButton(text="😐 Нормально", callback_data="mood_neutral"),
         InlineKeyboardButton(text="😡 Жалко", callback_data="mood_bad")]
    ])

pending_moods = {}

# --- ФУНКЦИИ БАЗЫ ---
def save_transaction(user_id, t_type, amount, category, comment, mood=None):
    today = datetime.now().strftime("%d.%m.%Y")
    cursor.execute('''
        INSERT INTO transactions (user_id, type, amount, category, date, comment, mood)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    ''', (user_id, t_type, amount, category, today, comment, mood))
    conn.commit()

def get_user_limit(user_id):
    cursor.execute("SELECT daily_limit FROM user_limits WHERE user_id = ?", (user_id,))
    res = cursor.fetchone()
    return res[0] if res else 1500

def set_user_limit(user_id, new_limit):
    cursor.execute("INSERT OR REPLACE INTO user_limits (user_id, daily_limit) VALUES (?, ?)", (user_id, new_limit))
    conn.commit()

def get_balance(user_id):
    cursor.execute("SELECT SUM(amount) FROM transactions WHERE user_id = ? AND type = 'Доход'", (user_id,))
    income = cursor.fetchone()[0] or 0
    cursor.execute("SELECT SUM(amount) FROM transactions WHERE user_id = ? AND type = 'Расход'", (user_id,))
    expense = cursor.fetchone()[0] or 0
    return income - expense

def get_category_expenses(user_id, days=30):
    cursor.execute("SELECT category, SUM(amount) FROM transactions WHERE user_id = ? AND type = 'Расход' AND date >= date('now', ?) AND category != 'Не указана' GROUP BY category", (user_id, f'-{days} days'))
    return cursor.fetchall()

def get_today_expenses(user_id):
    today = datetime.now().strftime("%d.%m.%Y")
    cursor.execute("SELECT SUM(amount) FROM transactions WHERE user_id = ? AND type = 'Расход' AND date = ?", (user_id, today))
    return cursor.fetchone()[0] or 0

def get_yesterday_expenses(user_id):
    yesterday = (datetime.now() - timedelta(days=1)).strftime("%d.%m.%Y")
    cursor.execute("SELECT SUM(amount) FROM transactions WHERE user_id = ? AND type = 'Расход' AND date = ?", (user_id, yesterday))
    return cursor.fetchone()[0] or 0

def get_avg_daily_expense(user_id):
    cursor.execute("SELECT AVG(amount) FROM (SELECT SUM(amount) as amount FROM transactions WHERE user_id = ? AND type = 'Расход' GROUP BY date ORDER BY date DESC LIMIT 7)", (user_id,))
    res = cursor.fetchone()[0]
    return int(res) if res else 0

def get_all_transactions_for_excel(user_id):
    cursor.execute("SELECT date, type, amount, category, comment, mood FROM transactions WHERE user_id = ? ORDER BY date DESC", (user_id,))
    return cursor.fetchall()

# --- ПАРСЕР ---
def parse_money(text):
    amounts = re.findall(r'\b\d+\b', text)
    if not amounts:
        return None, None, None, "Я не нашел сумму."
    amount = int(amounts[0])
    t_lower = text.lower()
    
    income_keywords = ['получил', 'зарплата', 'заработал', 'доход', 'пришло', 'перевели', 'аванс', 'продал', 'премия', 'выплата', 'подработка', 'халтура', 'нашел', 'вернули', 'расчет', 'гонорар', 'кешбэк', 'выиграл', 'дивиденды']
    if any(word in t_lower for word in income_keywords):
        return "Доход", amount, "Доход", None
        
    transport_keywords = ['такси', 'метро', 'автобус', 'заправка', 'бензин', 'парковка', 'машина', 'авто', 'заправил', 'топливо', 'газ', 'электричка', 'самолет', 'билет']
    if any(word in t_lower for word in transport_keywords):
        return "Расход", amount, "Транспорт", None
        
    food_keywords = ['кафе', 'ресторан', 'поужинал', 'обед', 'шаурма', 'кофе', 'пицца', 'продукты', 'магазин', 'суши', 'бургер', 'еда', 'завтрак', 'ужин', 'фастфуд', 'доставка', 'лавка', 'рынок']
    if any(word in t_lower for word in food_keywords):
        return "Расход", amount, "Еда", None
        
    home_keywords = ['квартира', 'аренда', 'коммуналка', 'свет', 'вода', 'газ', 'интернет', 'телефон', 'связь', 'хозяева', 'жкх', 'ремонт', 'дом', 'ипотека']
    if any(word in t_lower for word in home_keywords):
        return "Расход", amount, "Жилье", None
        
    fun_keywords = ['кино', 'бар', 'клуб', 'концерт', 'игры', 'плейстейшн', 'развлечения', 'караоке', 'вечеринка', 'боулинг', 'квест']
    if any(word in t_lower for word in fun_keywords):
        return "Расход", amount, "Развлечения", None
        
    shopping_keywords = ['купил', 'телефон', 'ноутбук', 'одёжда', 'штаны', 'куртка', 'кроссовки', 'ботинки', 'джинсы', 'техника', 'наушники', 'планшет', 'часы', 'сумка']
    if any(word in t_lower for word in shopping_keywords):
        return "Расход", amount, "Покупки", None
        
    health_keywords = ['врач', 'лекарство', 'аптека', 'больница', 'таблетки', 'зубы', 'стоматолог', 'анализы', 'массаж', 'косметолог']
    if any(word in t_lower for word in health_keywords):
        return "Расход", amount, "Здоровье", None
    
    if any(word in t_lower for word in ['потратил', 'расход', 'оплатил', 'отдал', 'снял', 'сходил']):
        return "Расход", amount, "Прочее", None
        
    return "Расход", amount, "Прочее", None

# --- ГРАФИК ДЛЯ TELEGRAM ---
def create_stats_chart(user_id):
    categories = get_category_expenses(user_id)
    if not categories:
        return None
    labels = [cat[0] for cat in categories]
    sizes = [cat[1] for cat in categories]
    fig, ax = plt.subplots(figsize=(7, 7), facecolor='#F8F9FA')
    colors = ['#1E5631', '#4C9A2A', '#77AC30', '#A4C639', '#C5E17A', '#E2EFDA']
    wedges, texts, autotexts = ax.pie(sizes, labels=labels, autopct='%1.0f%%', startangle=90, colors=colors, wedgeprops={'width': 0.4, 'edgecolor': 'white', 'linewidth': 2})
    plt.setp(autotexts, size=10, weight="bold", color="white")
    plt.setp(texts, size=11, weight="bold", color="#333333")
    total = sum(sizes)
    ax.text(0, 0, f"{total:,.0f} ₽", ha='center', va='center', fontsize=20, fontweight='bold', color='#333333')
    ax.set_title("💰 Мои расходы", fontsize=16, color='#2C3E50', fontweight='bold', pad=20)
    buf = io.BytesIO()
    plt.savefig(buf, format='png', dpi=150, bbox_inches='tight', facecolor='#F8F9FA')
    buf.seek(0)
    plt.close()
    return buf

# --- 🌟 ЭКСПОРТ В EXCEL (С ДАШБОРДОМ И ГРАФИКОМ) ---
def create_excel_report(user_id):
    data = get_all_transactions_for_excel(user_id)
    if not data:
        return None
    
    df = pd.DataFrame(data, columns=['Дата', 'Тип', 'Сумма', 'Категория', 'Комментарий', 'Настроение'])
    mood_map = {'mood_good': '😊 Классно', 'mood_neutral': '😐 Нормально', 'mood_bad': '😡 Жалко', None: '-'}
    df['Настроение'] = df['Настроение'].map(mood_map)
    
    # Подсчет итогов для дашборда
    total_income = df[df['Тип'] == 'Доход']['Сумма'].sum()
    total_expense = df[df['Тип'] == 'Расход']['Сумма'].sum()
    balance = total_income - total_expense
    
    # Группировка по категориям для графика
    cat_summary = df[df['Тип'] == 'Расход'].groupby('Категория')['Сумма'].sum().reset_index()
    cat_summary = cat_summary.sort_values('Сумма', ascending=False)
    
    # Создаем Excel файл в памяти
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # 1. Лист с данными
        df.to_excel(writer, index=False, sheet_name='Все операции')
        
        # 2. Лист с Дашбордом
        workbook = writer.book
        dashboard = workbook.create_sheet('Дашборд')
        
        # Стили
        title_font = Font(size=16, bold=True, color='1E5631')
        big_font = Font(size=22, bold=True, color='1E5631')
        green_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        
        # Название Дашборда
        dashboard['A1'] = "📊 МОЙ ФИНАНСОВЫЙ ОТЧЕТ"
        dashboard['A1'].font = title_font
        dashboard.merge_cells('A1:C1')
        dashboard['A1'].alignment = Alignment(horizontal='center')
        
        # Основные цифры
        dashboard['A3'] = "💰 Доходы:"
        dashboard['A4'] = "💸 Расходы:"
        dashboard['A5'] = "💎 Баланс:"
        
        dashboard['B3'] = f"{total_income:,.0f} ₽"
        dashboard['B4'] = f"{total_expense:,.0f} ₽"
        dashboard['B5'] = f"{balance:,.0f} ₽"
        
        dashboard['B3'].font = big_font
        dashboard['B4'].font = big_font
        dashboard['B5'].font = big_font
        
        # Красим ячейки
        for row in range(3, 6):
            dashboard[f'A{row}'].fill = green_fill
            dashboard[f'B{row}'].fill = green_fill
        
        # Добавляем столбчатую диаграмму (Расходы по категориям)
        if not cat_summary.empty:
            # Записываем данные для графика на лист
            start_row = 8
            dashboard[f'A{start_row}'] = "Категория"
            dashboard[f'B{start_row}'] = "Сумма (₽)"
            dashboard[f'A{start_row}'].font = Font(bold=True)
            dashboard[f'B{start_row}'].font = Font(bold=True)
            
            for i, row in cat_summary.iterrows():
                r = start_row + i + 1
                dashboard[f'A{r}'] = row['Категория']
                dashboard[f'B{r}'] = row['Сумма']
            
            # Создаем объект диаграммы
            chart = BarChart()
            chart.title = "Траты по категориям"
            chart.y_axis.title = "Сумма (₽)"
            chart.x_axis.title = "Категория"
            chart.width = 10
            chart.height = 6
            
            # Определяем данные для графика
            data_range = Reference(dashboard, min_col=2, min_row=start_row, max_row=start_row + len(cat_summary))
            categories_range = Reference(dashboard, min_col=1, min_row=start_row+1, max_row=start_row + len(cat_summary))
            
            chart.add_data(data_range, titles_from_data=True)
            chart.set_categories(categories_range)
            chart.style = 10 # Выбираем красивый стиль
            
            # Вставляем график на лист (справа от цифр)
            dashboard.add_chart(chart, "E3")
            
        # Настраиваем ширину колонок
        dashboard.column_dimensions['A'].width = 20
        dashboard.column_dimensions['B'].width = 20
        dashboard.column_dimensions['E'].width = 30
        
    output.seek(0)
    return output

# --- УТРЕННИЙ ОТЧЕТ ---
async def send_morning_report():
    print("🌅 Отправляю утренние отчеты...")
    cursor.execute("SELECT DISTINCT user_id FROM transactions")
    users = cursor.fetchall()
    for user_tuple in users:
        user_id = user_tuple[0]
        try:
            yesterday_spent = get_yesterday_expenses(user_id)
            daily_limit = get_user_limit(user_id)
            balance = get_balance(user_id)
            report = f"☀️ **Доброе утро!** Отчет за вчера:\n\n"
            if yesterday_spent == 0:
                report += "📭 Вчера не было записей о тратах."
            else:
                over_limit = yesterday_spent - daily_limit
                if over_limit > 0:
                    report += f"🚨 Перерасход на **{over_limit} ₽**.\n"
                else:
                    report += f"✅ Ты уложился в лимит.\n💸 Потрачено: **{yesterday_spent} ₽** из {daily_limit} ₽.\n\n"
            report += f"💰 Баланс: **{balance} ₽**"
            await bot.send_message(user_id, report)
        except Exception as e:
            print(f"Ошибка отправки: {e}")

# --- ОБРАБОТЧИК СООБЩЕНИЙ ---
@dp.message()
async def handle_message(message: types.Message):
    user_id = message.from_user.id
    text = message.text
    
    if text == "💰 Баланс":
        balance = get_balance(user_id)
        await message.answer(f"💰 Твой текущий баланс: **{balance} ₽**", reply_markup=get_main_keyboard())
        return
        
    elif text == "📊 Статистика":
        chart = create_stats_chart(user_id)
        if chart:
            await message.answer_photo(photo=types.BufferedInputFile(chart.read(), filename="stats.png"), caption="📊 Кольцевая диаграмма твоих трат", reply_markup=get_main_keyboard())
        else:
            await message.answer("📭 Пока нет расходов.", reply_markup=get_main_keyboard())
        return
        
    elif text == "🎯 Цель":
        await message.answer("📝 Отправь цель: `/цель 100000 6`", reply_markup=get_main_keyboard())
        return

    elif text == "⚙️ Мой лимит":
        curr_limit = get_user_limit(user_id)
        await message.answer(f"⚙️ Твой лимит: **{curr_limit} ₽/день**\n\nНапиши новую сумму цифрами.", reply_markup=get_main_keyboard())
        return

    elif text == "🧘 Тренер":
        avg_spend = get_avg_daily_expense(user_id)
        if avg_spend == 0:
            await message.answer("📭 Запиши больше трат, я дам разбор!", reply_markup=get_main_keyboard())
            return
        reply = "🧘 **Твой коуч на сегодня:**\n\n"
        reply += f"📊 Твоя норма в день: **{avg_spend} ₽**.\n"
        reply += f"💡 Совет: Старайся тратить на 15% меньше нормы.\n"
        await message.answer(reply, reply_markup=get_main_keyboard())
        return

    elif text == "📁 Скачать Excel":
        excel_file = create_excel_report(user_id)
        if excel_file:
            await message.answer_document(
                document=types.BufferedInputFile(excel_file.read(), filename="finbro_report.xlsx"),
                caption="📄 Твой отчет с графиком и дашбордом готов!",
                reply_markup=get_main_keyboard()
            )
        else:
            await message.answer("📭 У тебя пока нет записей для отчета.", reply_markup=get_main_keyboard())
        return

    text_lower = text.lower()
    
    if text_lower == "/start":
        await message.answer(
            "🤖 **Finbro PRO** — с Excel дашбордом!\n\n"
            "📝 Пиши траты: 'Такси 350'\n"
            "📁 Скачай Excel, внутри таблица + график!\n\n"
            "👇 Жми на кнопки!",
            reply_markup=get_main_keyboard()
        )
    
    elif text_lower.startswith("/цель"):
        parts = text_lower.split()
        if len(parts) < 3:
            await message.answer("❌ Формат: /цель [сумма] [месяцев].", reply_markup=get_main_keyboard())
            return
        try:
            target = int(parts[1])
            months = int(parts[2])
        except:
            await message.answer("❌ Укажи цифры.", reply_markup=get_main_keyboard())
            return
        current_balance = get_balance(user_id)
        avg_daily_spend = get_avg_daily_expense(user_id)
        needed_per_month = target / months
        reply = f"🎯 **Цель:** {target} ₽ за {months} мес.\nНужно: **{needed_per_month:,.0f} ₽/мес**.\n\n"
        if current_balance >= target:
            reply += "✅ У тебя уже есть эта сумма!"
        else:
            reply += f"📊 Твои средние траты в день: **{avg_daily_spend} ₽**\n"
        await message.answer(reply, reply_markup=get_main_keyboard())

    elif text.isdigit():
        new_limit = int(text)
        set_user_limit(user_id, new_limit)
        await message.answer(f"✅ Лимит обновлен до **{new_limit} ₽**!", reply_markup=get_main_keyboard())

    else:
        t_type, amount, category, error = parse_money(text)
        if error:
            await message.answer(f"❌ {error}", reply_markup=get_main_keyboard())
        else:
            if t_type == "Доход":
                save_transaction(user_id, t_type, amount, category, text, mood="mood_good")
                sign = "+"
                response = f"✅ Записал {t_type}!\nСумма: {sign}{amount} ₽\nКатегория: 💰 Доход\n📅 {datetime.now().strftime('%d.%m.%Y')}\n\n💸 Баланс обновлен!"
                await message.answer(response, reply_markup=get_main_keyboard())
            else:
                pending_moods[user_id] = (t_type, amount, category, text)
                today_spent = get_today_expenses(user_id)
                daily_limit = get_user_limit(user_id)
                remaining = daily_limit - today_spent
                response = (
                    f"✅ Записал {t_type}!\n"
                    f"Сумма: -{amount} ₽\n"
                    f"Категория: {category}\n"
                    f"📅 {datetime.now().strftime('%d.%m.%Y')}\n\n"
                    f"💸 Сегодня потрачено: **{today_spent} ₽**\n"
                    f"📉 Остаток лимита: **{remaining} ₽**\n\n"
                    f"👇 Выбери настроение от этой траты:"
                )
                await message.answer(response, reply_markup=get_mood_keyboard())

@dp.callback_query()
async def process_mood_callback(callback_query: types.CallbackQuery):
    user_id = callback_query.from_user.id
    if user_id not in pending_moods:
        await callback_query.answer("⏳ Время выбора настроения истекло.")
        await callback_query.message.delete()
        return
    t_type, amount, category, comment = pending_moods.pop(user_id)
    mood = callback_query.data
    save_transaction(user_id, t_type, amount, category, comment, mood=mood)
    mood_emoji = {"mood_good": "😊", "mood_neutral": "😐", "mood_bad": "😡"}
    emoji = mood_emoji.get(mood, "😐")
    await callback_query.answer(f"Отлично! Настроение {emoji} сохранено.")
    await callback_query.message.edit_text(
        text=f"✅ Запись завершена с настроением {emoji}!\nТы можешь скачать Excel-отчет через меню, чтобы увидеть свои эмоции по тратам."
    )

async def main():
    scheduler = AsyncIOScheduler()
    scheduler.add_job(send_morning_report, CronTrigger(hour=8, minute=30))
    scheduler.start()
    print("🚀 FINBRO PRO (EXCEL DASHBOARD) ЗАПУЩЕН!")
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())
